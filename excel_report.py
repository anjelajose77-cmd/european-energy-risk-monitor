from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
TITLE_FONT  = Font(bold=True, size=14, color="1F3864")
BREACH_FILL = PatternFill("solid", fgColor="F8CBAD")
OK_FILL     = PatternFill("solid", fgColor="C6E0B4")
MONEY = '#,##0'

def header_row(ws, row, labels):
    for col, label in enumerate(labels, start=1):
        cell = ws.cell(row=row, column=col, value =label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

def _autosize(ws):
    for column in ws.columns:
        width = max((len(str(c.value)) for c in column if c.value is not None), default=10)
        ws.column_dimensions[column[0].column_letter].width = width + 3

def build_report(latest, book, total_pnl, var_hist, var_param, scenarios, limits,out_path):
    wb = Workbook()

    ws = wb.active
    ws.title="Summary"
    ws["A1"] = "European Cross-Commodity Risk Monitor"
    ws["A1"].font = TITLE_FONT
    ws["A2"]= f"Daily Risk Report - {date.today(): %d %B %Y}"
    ws["A2"].font = Font(italic=True, color="808080")
    summary = [("Total P&L (EUR)", total_pnl), ("VAR - historical (1d, 95%)", var_hist), ("VaR — Parametric (1d, 95%)", var_param), ("Limit breaches", sum(1 for a in limits if a["status"] == "BREACH")),]
    for i, (label, val) in enumerate(summary, start=4):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=val).number_format = MONEY
    _autosize(ws)

    ws= wb.create_sheet("P&L by Leg")
    header_row(ws, 1, ["Leg", "Direction", "P&L (EUR)"])
    for r, row in enumerate(book.itertuples(), start = 2):
        ws.cell(row=r, column=1, value=row.leg)
        ws.cell(row=r, column=2, value=row.direction)
        ws.cell(row=r, column=3, value=row.pnl).number_format=MONEY
    ws.cell(row=ws.max_row+1, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=ws.max_row, column=3, value=total_pnl).number_format=MONEY
    _autosize(ws)

    ws=wb.create_sheet("VaR")
    header_row(ws, 1, ["Method", "1-day 95% VaR (EUR)"])
    ws.cell(row=2, column=1, value="Historical simulation")
    ws.cell(row=2, column=2, value=var_hist).number_format = MONEY
    ws.cell(row=3, column=1, value="Parametric (var-covar)")
    ws.cell(row=3, column=2, value=var_param).number_format = MONEY
    _autosize(ws)

    ws = wb.create_sheet("Stress & Scenarios")
    header_row(ws, 1, ["Scenario", "Total P&L (EUR)"])
    for r, (name, pnl) in enumerate(scenarios.items(), start=2):
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=pnl).number_format = MONEY
    _autosize(ws)

    ws = wb.create_sheet("Limits & Alerts")
    header_row(ws, 1, ["Item", "Exposure", "Limit", "% Used", "Status"])
    for r, a in enumerate(limits, start=2):
        ws.cell(row=r, column=1, value=a["item"])
        ws.cell(row=r, column=2, value=a["exposure"]).number_format = MONEY
        ws.cell(row=r, column=3, value=a["limit"]).number_format = MONEY
        ws.cell(row=r, column=4, value=round(a["pct_used"], 0))
        ws.cell(row=r, column=5, value=a["status"])
        fill = BREACH_FILL if a["status"] == "BREACH" else OK_FILL
        for col in range(1, 6):
            ws.cell(row=r, column=col).fill = fill
    _autosize(ws)

    wb.save(out_path)
    return out_path


